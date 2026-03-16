#!/usr/bin/env python3
"""
Published psi sites ∩ L1 overlap analysis.

Data sources:
  - BID-seq (Dai et al., 2023 Nat Biotechnol) GSE179798: HeLa WT mRNA psi sites
  - BACS-seq (Zhao et al., 2024 Nat Chem Biol) GSE241849: HeLa polyA+ RNA psi sites

Both provide genomic coordinates → bedtools intersect with L1 BED.
"""
import subprocess
import os
import sys
import tempfile
from collections import Counter

# ── Paths ──
BASE = "/qbio/junsoopablo/02_Projects/05_IsoTENT_002_L1"
L1_BED = f"{BASE}/reference/L1_TE_L1_family.bed"
OUT_DIR = f"{BASE}/analysis/01_exploration/topic_05_cellline/pus_enzyme_analysis"
PSI_DIR = f"{OUT_DIR}/psi_sites"
BEDTOOLS = "conda run -n research bedtools"

YOUNG_L1 = {"L1HS", "L1PA1", "L1PA2", "L1PA3"}


def run(cmd):
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"  WARN cmd failed: {cmd[:80]}\n  {result.stderr[:200]}", file=sys.stderr)
    return result.stdout


def parse_bidseq_wt(xlsx_path):
    """Parse BID-seq HeLa WT xlsx → list of (chr, pos, strand, motif, gene, seg, frac)."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Sheet1']

    sites = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'chr' and row[1] == 'pos':
            header_found = True
            continue
        if not header_found or row[0] is None:
            continue
        chrom = str(row[0])
        pos = int(row[1])
        strand = str(row[5]) if row[5] else '+'
        motif1 = str(row[10]) if row[10] else ''
        gene = str(row[2]) if row[2] else ''
        seg = str(row[4]).strip() if row[4] else ''
        frac = float(row[15]) if row[15] else 0.0
        sites.append({
            'chr': chrom, 'pos': pos, 'strand': strand,
            'motif': motif1, 'gene': gene, 'seg': seg, 'frac': frac
        })
    return sites


def parse_bacsseq_polya(xlsx_path):
    """Parse BACS-seq Supplementary Table 8 (polyA+ RNA psi) → list of sites."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Supplementary Table 8']

    sites = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'chr' and row[1] == 'pos':
            header_found = True
            continue
        if not header_found or row[0] is None:
            continue
        chrom = str(row[0])
        pos = int(row[1])
        strand = str(row[2]) if row[2] else '+'
        motif = str(row[3]) if row[3] else ''
        gene = str(row[4]) if row[4] else ''
        feature = str(row[5]) if row[5] else ''
        psi_level = float(row[15]) if row[15] else 0.0
        sites.append({
            'chr': chrom, 'pos': pos, 'strand': strand,
            'motif': motif, 'gene': gene, 'feature': feature, 'psi_level': psi_level
        })
    return sites


def sites_to_bed(sites, bed_path):
    """Write sites to BED6 file (chr, start, end, name, score, strand)."""
    with open(bed_path, 'w') as f:
        for i, s in enumerate(sites):
            # BED is 0-based half-open; pos is 1-based → start = pos-1, end = pos
            start = s['pos'] - 1
            end = s['pos']
            name = s.get('gene', f'site_{i}')
            f.write(f"{s['chr']}\t{start}\t{end}\t{name}\t0\t{s['strand']}\n")
    return bed_path


def intersect_sites_with_l1(bed_path):
    """Intersect site BED with L1 BED, return set of overlapping site indices and L1 info."""
    # Use -wa -wb to get both entries
    cmd = f"{BEDTOOLS} intersect -a {bed_path} -b {L1_BED} -wa -wb"
    result = run(cmd)

    overlaps = []
    seen_sites = set()
    for line in result.strip().split('\n'):
        if not line:
            continue
        parts = line.split('\t')
        # site BED: 6 cols, L1 BED: 5 cols
        site_key = (parts[0], int(parts[1]), int(parts[2]), parts[5])  # chr, start, end, strand
        l1_subfamily = parts[6 + 3]  # col 4 of L1 BED
        l1_id = parts[6 + 4]  # col 5 of L1 BED
        if site_key not in seen_sites:
            seen_sites.add(site_key)
            overlaps.append({
                'chr': parts[0], 'pos': int(parts[1]) + 1,
                'strand': parts[5], 'site_name': parts[3],
                'l1_subfamily': l1_subfamily, 'l1_id': l1_id
            })
    return overlaps


def analyze_psi_sites():
    print("=" * 60)
    print("Published Psi Sites ∩ L1 Overlap Analysis")
    print("=" * 60)

    # ── Parse datasets ──
    print("\n[1] Parsing BID-seq HeLa WT...")
    bidseq_sites = parse_bidseq_wt(f"{PSI_DIR}/BIDseq_HeLa_WT.xlsx")
    print(f"  {len(bidseq_sites)} psi sites")

    print("\n[2] Parsing BACS-seq polyA+ RNA...")
    bacsseq_sites = parse_bacsseq_polya(f"{PSI_DIR}/BACSseq_tables.xlsx")
    print(f"  {len(bacsseq_sites)} psi sites")

    # ── Create BED files ──
    print("\n[3] Creating BED files...")
    bid_bed = f"{PSI_DIR}/bidseq_hela_wt.bed"
    bacs_bed = f"{PSI_DIR}/bacsseq_polya.bed"
    sites_to_bed(bidseq_sites, bid_bed)
    sites_to_bed(bacsseq_sites, bacs_bed)

    # Sort BED files
    for bed in [bid_bed, bacs_bed]:
        sorted_bed = bed + '.sorted'
        run(f"sort -k1,1 -k2,2n {bed} > {sorted_bed} && mv {sorted_bed} {bed}")

    # ── Intersect with L1 ──
    results = {}
    for dataset_name, sites, bed_path in [
        ("BID-seq_HeLa_WT", bidseq_sites, bid_bed),
        ("BACS-seq_polyA+", bacsseq_sites, bacs_bed),
    ]:
        print(f"\n[4] Intersecting {dataset_name} with L1...")
        overlaps = intersect_sites_with_l1(bed_path)

        total = len(sites)
        in_l1 = len(overlaps)
        pct = in_l1 / total * 100 if total > 0 else 0

        # Subfamily breakdown
        subfam_counts = Counter(ov['l1_subfamily'] for ov in overlaps)
        young_count = sum(v for k, v in subfam_counts.items() if k in YOUNG_L1)
        ancient_count = sum(v for k, v in subfam_counts.items() if k not in YOUNG_L1)

        # Motif breakdown for L1-overlapping sites
        # Match back to original sites to get motifs
        overlap_positions = {(ov['chr'], ov['pos'], ov['strand']) for ov in overlaps}
        l1_motifs = Counter()
        non_l1_motifs = Counter()
        for s in sites:
            key = (s['chr'], s['pos'], s['strand'])
            if key in overlap_positions:
                if s['motif']:
                    l1_motifs[s['motif']] += 1
            else:
                if s['motif']:
                    non_l1_motifs[s['motif']] += 1

        # Feature breakdown for L1 sites (BACS-seq only)
        if 'feature' in sites[0]:
            l1_features = Counter()
            for s in sites:
                key = (s['chr'], s['pos'], s['strand'])
                if key in overlap_positions:
                    l1_features[s.get('feature', 'unknown')] += 1
        else:
            l1_features = Counter()

        # Enrichment: compare observed vs expected
        # Expected: ~17.5% of genome is L1
        l1_genome_frac = 0.1747
        expected_in_l1 = total * l1_genome_frac
        enrichment = in_l1 / expected_in_l1 if expected_in_l1 > 0 else 0

        # Fisher test
        from scipy.stats import fisher_exact
        # Compare: sites_in_L1 vs sites_not_in_L1, relative to genome L1 fraction
        # Use genome bp as denominator
        genome_bp = 3_088_286_401
        l1_bp = 539_634_158
        table = [[in_l1, total - in_l1], [l1_bp, genome_bp - l1_bp]]
        odds_ratio, p_value = fisher_exact(table, alternative='two-sided')

        print(f"  Total sites: {total}")
        print(f"  Sites in L1: {in_l1} ({pct:.1f}%)")
        print(f"  Expected (genome fraction): {expected_in_l1:.0f} ({l1_genome_frac*100:.1f}%)")
        print(f"  Enrichment: {enrichment:.2f}x")
        print(f"  Fisher OR: {odds_ratio:.4f}, p={p_value:.2e}")
        print(f"  Young L1: {young_count}, Ancient: {ancient_count}")
        if l1_motifs:
            print(f"  L1 motifs: {l1_motifs.most_common(10)}")
        if l1_features:
            print(f"  L1 features: {l1_features.most_common()}")

        # Show L1-overlapping sites detail
        if overlaps:
            print(f"\n  L1-overlapping sites detail:")
            for ov in overlaps[:20]:
                print(f"    {ov['chr']}:{ov['pos']} {ov['strand']} → {ov['l1_subfamily']} ({ov['l1_id']})")
            if len(overlaps) > 20:
                print(f"    ... and {len(overlaps) - 20} more")

        results[dataset_name] = {
            'total': total, 'in_l1': in_l1, 'pct': pct,
            'enrichment': enrichment, 'fisher_or': odds_ratio, 'fisher_p': p_value,
            'young': young_count, 'ancient': ancient_count,
            'subfamilies': subfam_counts, 'l1_motifs': l1_motifs,
            'non_l1_motifs': non_l1_motifs,
        }

    # ── Compare motif distribution: L1 vs non-L1 ──
    print("\n" + "=" * 60)
    print("Motif comparison: L1 vs non-L1 psi sites")
    print("=" * 60)
    for ds_name, r in results.items():
        print(f"\n  {ds_name}:")
        all_motifs = set(r['l1_motifs'].keys()) | set(r['non_l1_motifs'].keys())
        if not all_motifs:
            print("    No motif data available")
            continue
        print(f"    {'Motif':<8} {'L1':>5} {'non-L1':>7} {'L1%':>6} {'nonL1%':>7}")
        total_l1 = sum(r['l1_motifs'].values())
        total_non = sum(r['non_l1_motifs'].values())
        for motif in sorted(all_motifs):
            l1_n = r['l1_motifs'].get(motif, 0)
            non_n = r['non_l1_motifs'].get(motif, 0)
            l1_pct = l1_n / total_l1 * 100 if total_l1 > 0 else 0
            non_pct = non_n / total_non * 100 if total_non > 0 else 0
            if l1_n > 0 or non_n >= 5:
                print(f"    {motif:<8} {l1_n:>5} {non_n:>7} {l1_pct:>5.1f}% {non_pct:>6.1f}%")

    # ── Save results ──
    out_tsv = f"{OUT_DIR}/pus_dependent_psi_in_l1.tsv"
    with open(out_tsv, 'w') as f:
        f.write("dataset\ttotal_sites\tsites_in_L1\tpct_in_L1\tenrichment\t"
                "fisher_OR\tfisher_p\tyoung_L1\tancient_L1\ttop_subfamilies\ttop_L1_motifs\n")
        for ds_name, r in results.items():
            top_sub = "; ".join(f"{k}({v})" for k, v in r['subfamilies'].most_common(5))
            top_mot = "; ".join(f"{k}({v})" for k, v in r['l1_motifs'].most_common(5))
            f.write(f"{ds_name}\t{r['total']}\t{r['in_l1']}\t{r['pct']:.2f}\t"
                    f"{r['enrichment']:.3f}\t{r['fisher_or']:.4f}\t{r['fisher_p']:.2e}\t"
                    f"{r['young']}\t{r['ancient']}\t{top_sub}\t{top_mot}\n")
    print(f"\n  Results saved: {out_tsv}")

    # ── Also check if any BID-seq shControl-specific sites were lost ──
    # The shControl file represents sites still present after control shRNA treatment
    # Sites in WT but not in shControl could indicate off-target or noise
    print("\n[5] Checking BID-seq shControl overlap...")
    shctrl_sites = parse_bidseq_shcontrol(f"{PSI_DIR}/BIDseq_HeLa_shControl.xlsx")
    print(f"  shControl sites: {len(shctrl_sites)}")

    # Cross-reference
    wt_set = {(s['chr'], s['pos']) for s in bidseq_sites}
    ctrl_set = {(s['chr'], s['pos']) for s in shctrl_sites}
    shared = wt_set & ctrl_set
    wt_only = wt_set - ctrl_set
    ctrl_only = ctrl_set - wt_set
    print(f"  WT-only: {len(wt_only)}, Control-only: {len(ctrl_only)}, Shared: {len(shared)}")


def parse_bidseq_shcontrol(xlsx_path):
    """Parse BID-seq HeLa shControl."""
    import openpyxl
    wb = openpyxl.load_workbook(xlsx_path, read_only=True)
    ws = wb['Sheet1']

    sites = []
    header_found = False
    for row in ws.iter_rows(values_only=True):
        if row[0] == 'chr' and row[1] == 'pos':
            header_found = True
            continue
        if not header_found or row[0] is None:
            continue
        chrom = str(row[0])
        pos = int(row[1])
        strand = str(row[5]) if row[5] else '+'
        sites.append({'chr': chrom, 'pos': pos, 'strand': strand})
    return sites


if __name__ == "__main__":
    analyze_psi_sites()
